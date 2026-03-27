#!/usr/bin/env python3
"""Quartiersatlas ETL: Excel + GeoJSON → enriched GeoJSON for the interactive map."""

import json
import math
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = ROOT / "Quartiersatlas_2024_Daten.xlsx"
GEOJSON_INPUT = ROOT / "Sozialräume_WGS84_4326_0.geojson"
GEOJSON_OUTPUT = ROOT / "data" / "sozialraeume.geojson"

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
UNBEWOHNT_ID = "071009"
GEMEINSAM_IDS = ("032001", "033001")  # 033001 receives 032001's index values

# Indicators for composite indices: (excel_col_index, clean_key, invert)
# "invert" means higher raw value = LESS social need (so z-score is negated)
SOZIAL_INDICATORS = [
    (31, "arbeitslosenquote", False),
    (33, "sgb2_quote", False),
    (34, "kinderarmut", False),
    (35, "altersarmut", False),
    (36, "mindestsicherung", False),
    (44, "uebergang_gym", True),      # higher gymnasium rate = less need → invert
]

FLUKTUATION_INDICATORS = [
    (24, "fluktuationsrate", False),
]

# Raw indicators to include in output GeoJSON: (excel_col_index, property_key)
RAW_PROPERTIES = [
    (3,  "bevoelkerung"),
    (4,  "weiblich_pct"),
    (19, "jugendquotient"),
    (20, "altenquotient"),
    (21, "auslaender_pct"),
    (22, "migration_pct"),
    (23, "wanderungssaldo"),
    (24, "fluktuationsrate"),
    (25, "haushalte"),
    (26, "einpersonen_hh_pct"),
    (27, "hh_kinder_pct"),
    (28, "alleinerziehende_pct"),
    (29, "senioren_single_pct"),
    (30, "arbeitslose"),
    (31, "arbeitslosenquote_pct"),
    (32, "sgb2_personen"),
    (33, "sgb2_quote_pct"),
    (34, "kinderarmut_pct"),
    (35, "altersarmut_pct"),
    (36, "mindestsicherung_pct"),
    (37, "wohngeld_hh_pct"),
    (44, "uebergang_gym_pct"),
    (45, "wohnflaeche_m2_ew"),
    (46, "oeff_gef_whg_pct"),
    (47, "wohneigentum_pct"),
    (48, "flaeche_ha"),
    (49, "bev_dichte_km2"),
    (51, "gruenflaeche_pct"),
]


def read_excel():
    """Read Excel → list of dicts with column-index keys plus 'id', 'name', 'stadtbezirk'."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Profildaten"]

    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue
        rec = {i: row[i] for i in range(len(row))}
        rec["id"] = str(row[0])
        rec["name"] = row[2]
        rec["stadtbezirk"] = row[1]
        records.append(rec)
    return records


def calc_z_scores(records, col_idx, exclude_ids=None):
    """Calculate z-scores for one column. Returns dict {sozialraum_id: z_value}."""
    exclude = exclude_ids or set()
    vals = {}
    for r in records:
        if r["id"] in exclude or r[col_idx] is None:
            continue
        vals[r["id"]] = float(r[col_idx])

    if len(vals) < 2:
        return {}

    n = len(vals)
    mean = sum(vals.values()) / n
    variance = sum((v - mean) ** 2 for v in vals.values()) / n
    std = math.sqrt(variance) if variance > 0 else 1.0
    return {sid: (v - mean) / std for sid, v in vals.items()}


def classify_z(z):
    """Map a z-value to one of the five Typisierung labels."""
    if z is None:
        return None
    if z < -1.0:
        return "gering"
    if z < -0.5:
        return "eher gering"
    if z < 0.5:
        return "mittel"
    if z < 1.0:
        return "erhöht"
    return "hoch"


def calc_composite(z_dicts):
    """Average multiple z-score dicts into a single composite index dict."""
    all_ids = set()
    for zd in z_dicts:
        all_ids.update(zd.keys())

    result = {}
    for sid in all_ids:
        values = [zd[sid] for zd in z_dicts if sid in zd]
        if values:
            result[sid] = sum(values) / len(values)
    return result


def process_indices(records):
    """Calculate composite z-Werte for Sozial and Fluktuation indices."""
    exclude = {UNBEWOHNT_ID}

    sozial_z_list = []
    sozial_z_by_key = {}
    for col_idx, key, invert in SOZIAL_INDICATORS:
        z = calc_z_scores(records, col_idx, exclude)
        if invert:
            z = {sid: -v for sid, v in z.items()}
        sozial_z_list.append(z)
        sozial_z_by_key[key] = z

    flukt_z_list = []
    flukt_z_by_key = {}
    for col_idx, key, invert in FLUKTUATION_INDICATORS:
        z = calc_z_scores(records, col_idx, exclude)
        if invert:
            z = {sid: -v for sid, v in z.items()}
        flukt_z_list.append(z)
        flukt_z_by_key[key] = z

    idx_sozial = calc_composite(sozial_z_list)
    idx_flukt = calc_composite(flukt_z_list)

    # Gemeinsam typisiert: 033001 receives 032001's composite values
    src, tgt = GEMEINSAM_IDS
    for idx in [idx_sozial, idx_flukt]:
        if src in idx:
            idx[tgt] = idx[src]

    return idx_sozial, idx_flukt, sozial_z_by_key, flukt_z_by_key


def build_enriched_properties(records, idx_sozial, idx_flukt, sozial_z_by_key, flukt_z_by_key):
    """Build a dict {sozialraum_id: {property_key: value, ...}} for GeoJSON merge."""
    enriched = {}
    for rec in records:
        sid = rec["id"]
        props = {
            "name": rec["name"],
            "stadtbezirk": rec["stadtbezirk"],
            "unbewohnt": sid == UNBEWOHNT_ID,
        }

        for col_idx, key in RAW_PROPERTIES:
            props[key] = rec.get(col_idx)

        zs = idx_sozial.get(sid)
        zf = idx_flukt.get(sid)
        props["z_sozial"] = round(zs, 4) if zs is not None else None
        props["z_fluktuation"] = round(zf, 4) if zf is not None else None
        props["typ_sozial"] = classify_z(zs)
        props["typ_fluktuation"] = classify_z(zf)

        for key, z_dict in sozial_z_by_key.items():
            props[f"z_{key}"] = round(z_dict[sid], 4) if sid in z_dict else None
        for key, z_dict in flukt_z_by_key.items():
            props[f"z_{key}"] = round(z_dict[sid], 4) if sid in z_dict else None

        enriched[sid] = props
    return enriched


def merge_geojson(enriched):
    """Read source GeoJSON, merge enriched properties, write output."""
    with open(GEOJSON_INPUT, "r", encoding="utf-8") as f:
        geojson = json.load(f)

    matched = 0
    for feature in geojson["features"]:
        sid = feature["properties"]["SOZIALRAUM_ID"]
        if sid in enriched:
            feature["properties"].update(enriched[sid])
            matched += 1
        else:
            feature["properties"]["unbewohnt"] = True
            feature["properties"]["name"] = "Unbekannt"

    GEOJSON_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with open(GEOJSON_OUTPUT, "w", encoding="utf-8") as f:
        json.dump(geojson, f, ensure_ascii=False)

    print(f"Matched {matched}/{len(geojson['features'])} features")
    print(f"Output: {GEOJSON_OUTPUT} ({GEOJSON_OUTPUT.stat().st_size / 1024:.0f} KB)")


if __name__ == "__main__":
    print("Reading Excel...")
    records = read_excel()
    print(f"  {len(records)} records")

    print("Calculating z-Werte...")
    idx_sozial, idx_flukt, sozial_z_by_key, flukt_z_by_key = process_indices(records)
    print(f"  Sozial: {len(idx_sozial)} | Fluktuation: {len(idx_flukt)}")

    print("Building properties...")
    enriched = build_enriched_properties(records, idx_sozial, idx_flukt, sozial_z_by_key, flukt_z_by_key)

    print("Merging into GeoJSON...")
    merge_geojson(enriched)

    # Spot-check
    print("\nSpot-check:")
    for sid in ["011001", "071009", "032001", "033001"]:
        e = enriched.get(sid, {})
        print(f"  {sid} ({e.get('name', '?')}): "
              f"typ_s={e.get('typ_sozial')}, typ_f={e.get('typ_fluktuation')}, "
              f"z_s={e.get('z_sozial')}, z_f={e.get('z_fluktuation')}, "
              f"unbewohnt={e.get('unbewohnt')}")
